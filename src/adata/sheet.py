#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Modulo para manejar archivos Excel.

 * Funciona con openpyxl
 * Necesita xlrd

Hoja Excel virtual

"""

import os
import xlrd
from prettytable import PrettyTable
from os.path import join, dirname, abspath, isfile
from xlrd.sheet import ctype_text  
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Color, colors, NamedStyle
from openpyxl.drawing.image import Image
from openpyxl.chart import ScatterChart, Reference, Series


def is_number(s):
    """ True solo si es un valor numerico
    """
    if s is None: 
        return False
    try:
        float(s)
        return True
    except ValueError:
        return False
    except TypeError:
        return False
        print ("Warning: is_number.TypeError %s" % s)

class UnNamedStyle(object):
    def apply(self, cell):
        cell.font = self.font


STYLE = {
#    "normal": NamedStyle(name="normal"),
#    "BOLD": NamedStyle(name="bold"),
#    "RED": NamedStyle(name="red")
    "normal": UnNamedStyle(),
    "BOLD": UnNamedStyle(),
    "RED": UnNamedStyle()
}
STYLE["normal"].font = Font(name='Arial', size=11, bold=False)
STYLE["BOLD"].font = Font(name='Arial', size=11, bold=True)
STYLE["RED"].font = Font(name='Arial', size=11, color=colors.RED)









class Sheet(object):
    """ Una hoja de calculo generica.
        Incluye formato basico
    """
    def __init__(self, name):
        self.name = name
        self.cols = []
        self.rows = []
        self.row = 0
        self.row_style = {}
        self.col_width = {}
        self.col_number_format = {}

    def write(self, row=None, style=None):
        if row is None:
            self.rows.append([])
        if type(row) is list:
            while len(row)<len(self.cols): row.append(None)
            self.rows.append(row)
        elif type(row) is dict:
            r = []
            c = 0
            for col in self.cols: # PROBLEMA: Si no se declaran todas las columnas no llega mas lejos
                c+=1 #1 -> A
                letter = get_column_letter(c) #A
                if row.has_key(c): r.append(row[c]) # Si c es numero
                elif row.has_key(col): r.append(row[col]) # si es clave
                elif row.has_key(letter): r.append(row[letter]) #si es letra
                else: r.append(None)
            self.rows.append(r)
        else:
            print("Error en Sheet.add_row %s" % row)

        if style is not None:

            i = len(self.rows)-1
            self.row_style[i]=STYLE[style]

    def get_col(self, x, y):
        #print "lista_columna %s %s " % (x, y)
        lista = []
        while True:
            y+=1
            if y>=len(self.rows): break
            if x>=len(self.rows[y]): break
            v = self.rows[y][x]
            if v is None or v=='': break
            lista.append(v)
        #print " %s" % lista
        return lista


    def get_row(self, x, y):
        if y>=len(self.rows): return
        #print "lista_fila %s %s " % (x, y)
        lista = []
        while True:
            x+=1
            if x>=len(self.rows[y]): break
            v = self.rows[y][x]
            if v is None or v=='': break 
            lista.append(v)    
        #print " %s" % lista
        return lista
            

    def dump(self):
        print("Hoja "+ self.name)
        print ('Hoja "%s" (%s columnas, %s filas)' % (self.name, len(self.cols), len(self.rows) ))
        pt = PrettyTable(self.cols) 
        
        for col in self.cols: pt.align[col] = "l"
        for row in self.rows:
            if row==self.cols: continue
            pt.add_row(row)
        print(pt.get_string())


    def get_value_position(self, value):
        """ Posicion (x,y) de la primera celda que sea este valor
        """
        y=0
        for row in self.rows:
            x=0
            for cell in row:
                if cell==value: return x,y
                x+=1
            y+=1
        return None
    
    def __iter__(self):
        # para Leer_XLS.listar

        cols = len(self.cols)
        for row in self.rows:
            #if row==self.cols: continue
            yield { self.cols[x]: row[x] for x in range(0, cols) }


    def write_to(self, ws, format=True):
        for c,w in self.col_width.items(): 
            if c in self.cols:
                letter = get_column_letter(self.cols.index(c))
            else:
                letter = get_column_letter(c) if is_number(c) else c
            ws.column_dimensions[letter].width = w

        for y in range(0, len(self.rows)):
            col = self.rows[y]
            for x in range(0, len(col)):
                letter = get_column_letter(x+1)
                cell = ws.cell(row = y+1, column = x+1)
                try:
                    cell.value = col[x]
                    if is_number(cell.value) and letter in self.col_number_format.keys():
                        #cell.number_format = '0'
                        cell.set_explicit_value(col[x], data_type="n")    
                    else:
                        cell.set_explicit_value(str(col[x]), data_type="s")
                except:
                    cell.value = "ERR %s" % repr(col[x])
                
                if format:
                    if y in self.row_style.keys():
                        #cell.style = self.row_style[y]
                        self.row_style[y].apply(cell)

                    else:
                        #cell.style = STYLE["normal"]
                        STYLE["normal"].apply(cell)


                    if is_number(cell.value):
                        if x in self.col_number_format.keys():
                            cell.number_format = self.col_number_format[x] #'0.00E+00'
                        elif letter in self.col_number_format.keys():
                            cell.number_format = self.col_number_format[letter] #'0.00E+00'
                












           

class Excel(object):
    '''Spreadsheet
    
    Load files:
        -XLS - xlrd
        -XLSX - openpyxl
    
    Save XLSX files

    '''

    def __init__(self, archivo = None):
        self.sheet = []
        self.archivo = archivo
        if archivo is None:
            self.wb = Workbook()
            return
        if not isfile(archivo):
            self.wb = Workbook()
            return

        self.filename, self.extension = os.path.splitext(archivo)
        print("Excel: Cargando %s desde %s" % (self.extension[1:].upper(), archivo))
       
        ext = self.extension.lower()
        if ext==".xls": self.wb = self._read_xls()
        elif ext==".xlsx": self.wb = self._read_xlsx()
        else: print("Excel: Extensión %s no soportada" % ext)
        

    def dump(self):
        print("Excel %s con %s hojas" % (self.archivo, len(self.sheet)))
        for sheet in self.sheet:
            sheet.dump()

        
    def _read_xls(self):
        w = xlrd.open_workbook(self.archivo)
        self.sheet_names = w.sheet_names()
        for n in range(0, len(self.sheet_names)):
            sheet = Sheet(self.sheet_names[n])
            print("Excel: Hoja %s - %s" % (n, sheet.name))
            s = w.sheet_by_index(n)
            sheet.cols = [ cell_obj.value for idx, cell_obj in enumerate(s.row(0))]
            sheet.rows = [[ cell.value for c, cell in enumerate(r)] for r in s.get_rows() ]
            self.sheet.append(sheet)
        return w

    def _read_xlsx(self):
        w = load_workbook(filename = self.archivo, data_only=True)
        self.sheet_names = list(w.sheetnames)
        for n in range(0, len(self.sheet_names)):
            sheet = Sheet(self.sheet_names[n])
            s = w.worksheets[n]
            sheet.cols = [ cell.value for cell in s[1]]
            sheet.rows = [[ cell.value for cell in row] for row in s.iter_rows() ]
            self.sheet.append(sheet)
        return w
    
    def new_sheet(self, name):
        s = Sheet(name)
        self.sheet.append(s)
        #self.sheet_names.append(name)
        return s

    def create_sheets(self, format=True):
        if len(self.sheet)==0:
            print("No hay hojas!")
            return
        wb = self.wb
        ws = wb.active
        ws.title = self.sheet[0].name
        for n in range(0, len(self.sheet)):
            s = self.sheet[n]
            if n>0: ws = wb.create_sheet(s.name)
            s.write_to(ws, format)

    def get_cell(self, name, row, column):
        return self.wb[name].cell(row = row, column = column)

    def set_cell_fill(self, name, row, column, color):
        self.get_cell(name, row, column).fill = PatternFill(start_color=color, end_color=color, fill_type='solid')

    def save(self, archivo=None, create=True):
        if archivo is None:
            archivo = self.archivo
        if create: self.create_sheets()
        #image = Image('glaslogo.png')
        #print image
        #ws.add_image(image, 'A15')
        print("Escribiendo archivo: %s " % archivo)
        self.wb.save( archivo )


"""
z = Excel('test.xls')
z.dump()
z.save("salida.xlsx")
print z.sheet[0].get_value_position("150121")

x = Excel('salida.xlsx')
x.dump()

s = x.sheet[0]
s.add_row([1,2,3])
s.add_row({u'USUARIO ALTA':"YO",u'REEX':3})
x.dump()

for line in x.sheet[0]:
    print line




            for c,w in s.col_width.iteritems(): ws.column_dimensions[get_column_letter(c+1)].width = w

            for y in range(0, len(s.rows)):
                col = s.rows[y]
                for x in range(0, len(col)):
                    cell = ws.cell(row = y+1, column = x+1)
                    cell.value = col[x]
                    if y in s.row_style.keys():
                        cell.style = s.row_style[y]
                    if x>0 and x in s.col_number_format.keys():
                        cell.number_format = s.col_number_format[x] #'0.00E+00'

"""